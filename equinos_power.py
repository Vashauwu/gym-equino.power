import os
from datetime import datetime
from threading import Lock

from telegram import Update
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ContextTypes,
    ConversationHandler, filters
)

from openpyxl import Workbook, load_workbook

TOKEN = os.getenv("8512404051:AAHfjDrMKwlfkrN9DSTFcXz_scsnNdRg8AQ")  # ponlo como variable de entorno
EXCEL_FILE = "asistencias_gym.xlsx"
PHOTO_DIR = "evidencias"

WAITING_PHOTO = 1
excel_lock = Lock()


def init_excel():
    """Crea el Excel si no existe, con encabezados."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"
        ws.append([
            "timestamp", "user_id", "username",
            "photo_path", "message_id"
        ])
        wb.save(EXCEL_FILE)


def append_row(row):
    """Escribe una fila en Excel de forma segura."""
    with excel_lock:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Asistencias"]
        ws.append(row)
        wb.save(EXCEL_FILE)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Hola 👋\nUsa /gym para registrar asistencia con foto."
    )


async def gym_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Perfecto ✅\nAhora envíame la foto de evidencia.")
    return WAITING_PHOTO


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.photo:
        await update.message.reply_text("Porfa manda una foto (no archivo).")
        return WAITING_PHOTO

    user = update.effective_user
    ts = datetime.now().isoformat(timespec="seconds")

    os.makedirs(PHOTO_DIR, exist_ok=True)

    # Telegram manda varias resoluciones; usamos la más grande
    photo = update.message.photo[-1]
    file = await context.bot.get_file(photo.file_id)

    # Nombre de archivo único
    filename = f"{ts.replace(':','-')}_uid{user.id}_mid{update.message.message_id}.jpg"
    path = os.path.join(PHOTO_DIR, filename)

    await file.download_to_drive(path)

    row = [
        ts,
        user.id,
        user.username or "",
        path,
        update.message.message_id
    ]
    append_row(row)

    await update.message.reply_text("✅ Asistencia registrada. ¡Gracias!")
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Cancelado.")
    return ConversationHandler.END


def main():
    if not TOKEN:
        raise RuntimeError("Falta TELEGRAM_BOT_TOKEN en variable de entorno")

    init_excel()

    app = Application.builder().token(TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("gym", gym_cmd)],
        states={
            WAITING_PHOTO: [MessageHandler(filters.PHOTO, handle_photo)]
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(conv)

    print("Bot corriendo...")
    app.run_polling()


if name == "main":
    main()